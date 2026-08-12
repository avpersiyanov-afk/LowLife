# -*- coding: utf-8 -*-
"""
Запись таблицы адресов СКС в xlsx через openpyxl.

Отдельный модуль (не часть route_addressing.py) намеренно: openpyxl —
CPython-only зависимость, а route_addressing.py используется и с
IronPython2-кнопок (RenumberAddresses, RenumberSkudAddresses и т.д.),
которым эта зависимость не нужна и не должна быть обязательной для
запуска. Импортируется только из кнопок, объявивших `#! python3`
(движок CPython) в script.py — сейчас это ExportAddressesToExcel.

openpyxl импортируется внутри функции, а не на уровне модуля: так
IronPython2-скрипты, случайно затронувшие этот файл (например при
рефакторинге lib/), не падают на импорте отсутствующего пакета.
"""

COLUMNS = [
    (u"ID", "id"),
    (u"Категория", "category"),
    (u"X, мм", "x_mm"),
    (u"Y, мм", "y_mm"),
    (u"Адрес", "addr"),
    (u"Предыдущий адрес", "addr_prev"),
    (u"Тип прокладки кабеля", "cable_type"),
]


def export_addressing_to_excel(rows, file_path):
    """
    rows — список dict с ключами из COLUMNS (второй элемент каждого
    кортежа). Пишет один лист с жирным заголовком и автошириной колонок.
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = u"Адреса СКС"

    ws.append([label for label, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(key, u"") for _, key in COLUMNS])

    for col_idx, (label, key) in enumerate(COLUMNS, start=1):
        max_len = len(label)
        for row in rows:
            max_len = max(max_len, len(str(row.get(key, u""))))
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(file_path)
